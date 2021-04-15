from openpyxl import Workbook
from openpyxl import load_workbook
try:
    wb = load_workbook("daten/datei3.xlsx")
    ws=wb.active
except:
    print("Fehler beim Einlesen der Datei")
for col in ws.iter_cols(min_col=2, max_col=4, min_row=2, max_row=4, values_only=False):
    for cell in col:
        print(cell.value)
