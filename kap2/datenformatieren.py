from openpyxl import Workbook
import datetime
wb = Workbook()
ws = wb.active
ws["a1"] = datetime.datetime(2020, 7, 21)
ws["a2"] = 1
ws["a3"] = 3.14
ws["a4"] = "String"
ws["a5"] = True
print(ws['a1'].number_format)
print(ws['a1'].is_date)

print(ws['a2'].number_format)
print(ws['a2'].is_date)
print(ws['a3'].number_format)
print(ws['a4'].number_format)
print(ws['a5'].number_format)
print(ws['a6'].number_format)
try:
    wb.save("daten/datei4.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")

