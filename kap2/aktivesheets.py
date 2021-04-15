from openpyxl import Workbook
wb = Workbook()
ws2 = wb.create_sheet()
ws3 = wb.create_sheet()
print(wb.sheetnames)
print(wb._active_sheet_index)
wb._active_sheet_index=1
print(wb._active_sheet_index)

