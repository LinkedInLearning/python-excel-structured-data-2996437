from openpyxl import Workbook
from openpyxl import load_workbook

try:
    wb = load_workbook("daten/datei7.xlsx")
    ws=wb.active
except:
    print("Fehler beim Einlesen der Datei")

ws.unmerge_cells('b2:D2')
try:
    wb.save("daten/datei8.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")