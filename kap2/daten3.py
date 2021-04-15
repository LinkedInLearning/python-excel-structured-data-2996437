from openpyxl import Workbook
from openpyxl import load_workbook
try:
    wb = load_workbook("daten/datei3.xlsx")
    ws=wb.active
except:
    print("Fehler beim Einlesen der Datei")
for row in ws.iter_rows(min_row=1, max_col=7, max_row=4):
    for cell in row:
        print(cell.value)
