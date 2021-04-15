from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active
ws1.title="Sheet1"
ws2 = wb.create_sheet("Sheet2",-1)
ws3 = wb.create_sheet("Sheet3",0)
try:
    wb.save("daten/datei1.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")

