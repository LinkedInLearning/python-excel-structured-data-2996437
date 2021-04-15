from openpyxl import Workbook
from openpyxl import load_workbook
try:
    wb = load_workbook("daten/datei1.xlsx")
    print(wb.sheetnames)
except:
    print("Fehler beim Einlesen der Datei")


