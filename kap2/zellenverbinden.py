from openpyxl import Workbook
from openpyxl import load_workbook

try:
    wb = load_workbook("daten/datei3.xlsx")
    ws=wb.active
except:
    print("Fehler beim Einlesen der Datei")
ws.merge_cells('b2:D2')

try:
    wb.save("daten/datei7.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")