from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE

try:
    wb = load_workbook("daten/datei3.xlsx")
    ws=wb.active
except:
    print("Fehler beim Einlesen der Datei")
anzahlZeilen=0
for row in ws.values:
    anzahlZeilen+=1
print(anzahlZeilen)
print("sum" in FORMULAE)
ws.cell(row=anzahlZeilen+1, column=2, value="=sum(b2:b15)")    
try:
    wb.save("daten/datei5.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")