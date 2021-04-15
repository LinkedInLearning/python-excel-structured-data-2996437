from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["a1"] = 1
ws.append([2,3,4,5])
zellbereich = ws["c3":"e5"]
zellbereich[0][1].value=42
print(zellbereich)
colG = ws["G"]
for x in colG:
    x.value=777
col_range = ws['H:J']
row_range = ws[5:10]
ws.cell(row=6, column=2, value=10)
for x in range(10,20):
    for y in range(10,30):
        ws.cell(row=x, column=y, value=x*y)
try:
    wb.save("daten/datei2.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")

