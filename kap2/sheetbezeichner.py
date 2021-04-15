from openpyxl import Workbook
wb = Workbook()
print(wb.sheetnames)
ws2 = wb.create_sheet()
wb.create_sheet()

print(wb.sheetnames)
print(wb.active.title)
wb.active.title="Neu"
print(wb.sheetnames)