from openpyxl import Workbook
wb = Workbook()
#print(wb.get_sheet_names())
print(wb.sheetnames)
ws = wb.active
print(wb._active_sheet_index)
wb.create_sheet("Neu")
print(wb.sheetnames)
wb.move_sheet("Sheet",1)
print(wb.sheetnames)
wb.remove(wb["Sheet"])
print(wb.sheetnames)
