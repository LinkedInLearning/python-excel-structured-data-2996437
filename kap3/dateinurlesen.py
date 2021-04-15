from openpyxl import Workbook
from openpyxl import load_workbook
try:
    wb = load_workbook("daten/datei1.xlsx",read_only=True)
    print(wb.sheetnames)
except:
    print("Fehler beim Einlesen der Datei")

try:
    wb.save("daten/datei2.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")
