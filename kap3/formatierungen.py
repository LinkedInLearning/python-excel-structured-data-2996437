from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
wb = Workbook()
ws = wb.active

for x in range(1,10):
    for y in range(1,5):
        ws.cell(row=x, column=y, value=x*y)

ft = Font(color="FF0000")
ws["a1"].font = ft
ws["d4"].font = ft
ws["c3"].font = Font(color="00A36F", italic=True, size=42)
try:
    wb.save("daten/dateimitstyles.xlsx")
except:
    print("Fehler beim Schreiben der Exceldatei")
